"""
LaTeX-style index formatter
"""
from typing import List, Tuple
from datetime import datetime


class IndexFormatter:
    def __init__(self):
        pass

    @staticmethod
    def normalize_first_letter(text: str) -> str:
        """
        Get the first letter for grouping
        Numbers and special characters become "#"
        Letters A-Z remain as uppercase letters
        """
        if not text or len(text) == 0:
            return '#'

        first_char = text[0].upper()

        # Check if it's a letter A-Z
        if first_char.isalpha() and first_char.isupper():
            return first_char

        # Everything else (numbers, special characters) becomes "#"
        return '#'

    def format_latex_style(self, entries: List[Tuple[str, List[str]]],
                           metadata: dict = None) -> str:
        """
        Format index entries in LaTeX style
        """
        if metadata is None:
            metadata = {'index_name': 'Index', 'books': [], 'custom_properties': []}

        output = []

        # Add header
        output.append(f"% {metadata['index_name']}")
        output.append(f"% Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        output.append("")

        # Front page
        output.append("\\begin{titlepage}")
        output.append("\\centering")
        output.append("")
        output.append(f"\\vspace*{{2cm}}")
        output.append(f"{{\\Huge \\textbf{{{self.escape_latex(metadata['index_name'])}}}\\par}}")
        output.append("\\vspace{2cm}")
        output.append("")

        # Books
        if metadata['books']:
            output.append("\\textbf{Books:}\\\\[0.5cm]")
            for book in metadata['books']:
                page_info = f" ({book['page_count']} pages)" if book['page_count'] else ""
                output.append(f"Book {book['book_number']}: {self.escape_latex(book['book_name'])}{page_info}\\\\")
            output.append("\\vspace{1cm}")

        # Custom properties
        if metadata['custom_properties']:
            output.append("\\textbf{Properties:}\\\\[0.5cm]")
            for prop in metadata['custom_properties']:
                output.append(f"{self.escape_latex(prop['name'])}: {self.escape_latex(prop['value'])}\\\\")
            output.append("\\vspace{1cm}")

        output.append("\\vfill")
        output.append(f"{{\\small Generated: {datetime.now().strftime('%Y-%m-%d')}}}")
        output.append("\\end{titlepage}")
        output.append("")

        if not entries:
            output.append("% Empty index")
            return "\n".join(output)

        output.append("\\begin{theindex}")
        output.append("")

        # Sort entries with "#" (special chars/numbers) at the end
        def sort_key(item):
            term = item[0]
            first_letter = self.normalize_first_letter(term)
            return (first_letter != '#', term.lower())

        entries = sorted(entries, key=sort_key)

        current_letter = None

        for term, references in entries:
            # Get first letter (normalize special chars and numbers to "#")
            first_letter = self.normalize_first_letter(term)

            # Add letter heading if new letter
            if first_letter != current_letter:
                if current_letter is not None:
                    output.append("")  # Blank line between letter sections
                output.append(f"  \\indexspace")
                output.append(f"  \\textbf{{{first_letter}}}")
                output.append("")
                current_letter = first_letter

            # Format references
            ref_str = ", ".join(references)

            # Escape special LaTeX characters in term
            escaped_term = self.escape_latex(term)

            output.append(f"  \\item {escaped_term}, {ref_str}")

        output.append("")
        output.append("\\end{theindex}")

        return "\n".join(output)

    def format_plain_text(self, entries: List[Tuple[str, List[str]]],
                          metadata: dict = None) -> str:
        """
        Format index entries as plain text with letter headings
        """
        if metadata is None:
            metadata = {'index_name': 'Index', 'books': [], 'custom_properties': []}

        output = []

        # Front page
        output.append("=" * 60)
        output.append(metadata['index_name'].center(60))
        output.append("=" * 60)
        output.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        output.append("=" * 60)
        output.append("")

        # Books section
        if metadata['books']:
            output.append("Books:")
            output.append("-" * 60)
            for book in metadata['books']:
                page_info = f" ({book['page_count']} pages)" if book['page_count'] else ""
                output.append(f"  Book {book['book_number']}: {book['book_name']}{page_info}")
            output.append("")

        # Custom properties section
        if metadata['custom_properties']:
            output.append("Properties:")
            output.append("-" * 60)
            for prop in metadata['custom_properties']:
                output.append(f"  {prop['name']}: {prop['value']}")
            output.append("")

        output.append("=" * 60)
        output.append("INDEX".center(60))
        output.append("=" * 60)
        output.append("")

        if not entries:
            output.append("Empty index")
            output.append("")
            output.append("=" * 60)
            output.append(f"Created with Index it!  •  {metadata['index_name']}".center(60))
            output.append("=" * 60)
            return "\n".join(output)

        # Sort entries with "#" (special chars/numbers) at the end
        def sort_key(item):
            term = item[0]
            first_letter = self.normalize_first_letter(term)
            return (first_letter != '#', term.lower())

        entries = sorted(entries, key=sort_key)

        current_letter = None

        for term, references in entries:
            # Get first letter (normalize special chars and numbers to "#")
            first_letter = self.normalize_first_letter(term)

            # Add letter heading if new letter
            if first_letter != current_letter:
                if current_letter is not None:
                    output.append("")  # Blank line between letter sections
                output.append(f"{first_letter}")
                output.append("-" * 40)
                current_letter = first_letter

            # Format references
            ref_str = ", ".join(references)

            output.append(f"  {term}: {ref_str}")

        output.append("")
        output.append("=" * 60)
        output.append(f"Total entries: {len(entries)}")
        output.append("=" * 60)
        output.append("")
        output.append(f"Created with Index it!  •  {metadata['index_name']}".center(60))
        output.append("=" * 60)

        return "\n".join(output)

    def format_markdown(self, entries: List[Tuple[str, List[str]]],
                        metadata: dict = None) -> str:
        """
        Format index entries as Markdown
        """
        if metadata is None:
            metadata = {'index_name': 'Index', 'books': [], 'custom_properties': []}

        output = []

        # Front page
        output.append(f"# {metadata['index_name']}")
        output.append("")
        output.append(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
        output.append("")

        # Books
        if metadata['books']:
            output.append("## Books")
            output.append("")
            for book in metadata['books']:
                page_info = f" ({book['page_count']} pages)" if book['page_count'] else ""
                output.append(f"- **Book {book['book_number']}**: {book['book_name']}{page_info}")
            output.append("")

        # Custom properties
        if metadata['custom_properties']:
            output.append("## Properties")
            output.append("")
            for prop in metadata['custom_properties']:
                output.append(f"- **{prop['name']}**: {prop['value']}")
            output.append("")

        output.append("---")
        output.append("")
        output.append("# Index")
        output.append("")

        if not entries:
            output.append("*Empty index*")
            output.append("")
            output.append("---")
            output.append(f"*Created with Index it!  •  {metadata['index_name']}*")
            return "\n".join(output)

        # Sort entries with "#" (special chars/numbers) at the end
        def sort_key(item):
            term = item[0]
            first_letter = self.normalize_first_letter(term)
            return (first_letter != '#', term.lower())

        entries = sorted(entries, key=sort_key)

        current_letter = None

        for term, references in entries:
            # Get first letter (normalize special chars and numbers to "#")
            first_letter = self.normalize_first_letter(term)

            # Add letter heading if new letter
            if first_letter != current_letter:
                if current_letter is not None:
                    output.append("")  # Blank line between letter sections
                output.append(f"## {first_letter}")
                output.append("")
                current_letter = first_letter

            # Format references
            ref_str = ", ".join(references)

            output.append(f"- **{term}**: {ref_str}")

        output.append("")
        output.append(f"---")
        output.append(f"*Total entries: {len(entries)}*")
        output.append("")
        output.append(f"*Created with Index it!  •  {metadata['index_name']}*")

        return "\n".join(output)

    def format_pdf(self, entries: List[Tuple[str, List[str]]],
                   output_path: str, metadata: dict = None) -> bool:
        """
        Format index entries as a 4-column landscape PDF with continuous flow
        Returns True if successful
        """
        if metadata is None:
            metadata = {'index_name': 'Index', 'books': [], 'custom_properties': []}

        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, KeepTogether, Table, TableStyle, PageBreak, Image, FrameBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_LEFT, TA_CENTER
            from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
            from reportlab.platypus.frames import Frame
            import os

            if not entries:
                return False

            # Sort entries with "#" (special chars/numbers) at the end
            def sort_key(item):
                term = item[0]
                first_letter = self.normalize_first_letter(term)
                # Put "#" at end, otherwise sort alphabetically
                return (first_letter == '#', term.lower())

            entries = sorted(entries, key=sort_key)

            # Define page number footer function with index name
            def add_footer(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica', 9)
                page_num = canvas.getPageNumber()

                # Page number on right
                canvas.drawRightString(landscape(letter)[0] - 0.5 * inch, 0.3 * inch, str(page_num))

                # Index name and "Created with Index it!" on left
                footer_text = f"{metadata['index_name']}  •  Created with Index it!"
                canvas.drawString(0.5 * inch, 0.3 * inch, footer_text)

                canvas.restoreState()

            # Create PDF with landscape orientation
            doc = BaseDocTemplate(
                output_path,
                pagesize=landscape(letter),
                rightMargin=0.5 * inch,
                leftMargin=0.5 * inch,
                topMargin=0.75 * inch,
                bottomMargin=0.5 * inch
            )

            # Calculate frame dimensions for 4 columns
            page_width, page_height = landscape(letter)
            frame_width = (page_width - 1 * inch - 3 * 0.15 * inch) / 4  # subtract margins and gaps
            frame_height = page_height - 1.25 * inch  # subtract top and bottom margins

            # Create 4 frames (columns)
            frames = []
            for i in range(4):
                x = 0.5 * inch + i * (frame_width + 0.15 * inch)
                frame = Frame(
                    x, 0.5 * inch,
                    frame_width, frame_height,
                    leftPadding=6, rightPadding=6,
                    topPadding=6, bottomPadding=6,
                    showBoundary=0
                )
                frames.append(frame)

            # Create page template with 4 columns and footer
            template = PageTemplate(id='FourCol', frames=frames, onPage=add_footer)
            doc.addPageTemplates([template])

            # Styles
            styles = getSampleStyleSheet()

            # Get accent color from metadata (default to pink if not provided)
            accent_color = metadata.get('color_scheme', '#f2849e')

            # Title style for front page
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Title'],
                fontSize=24,
                textColor=colors.HexColor(accent_color),
                spaceAfter=12,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            # Front page heading style
            front_heading_style = ParagraphStyle(
                'FrontHeadingStyle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor(accent_color),
                spaceAfter=8,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )

            # Front page content style
            front_content_style = ParagraphStyle(
                'FrontContentStyle',
                parent=styles['Normal'],
                fontSize=11,
                spaceAfter=6,
                leftIndent=20
            )

            # Letter heading style (larger and bold)
            heading_style = ParagraphStyle(
                'LetterHeading',
                parent=styles['Heading1'],
                fontSize=14,
                textColor=colors.HexColor(accent_color),
                spaceAfter=8,
                spaceBefore=12,
                fontName='Helvetica-Bold',
                keepWithNext=True
            )

            # Entry style (term in bold, references in normal)
            entry_style = ParagraphStyle(
                'EntryStyle',
                parent=styles['Normal'],
                fontSize=9,
                leftIndent=10,
                spaceAfter=4,
                leading=11
            )

            # Build content as a continuous flow
            story = []

            # Column 1: Logo and title
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'indexit-logo.jpg')
            if os.path.exists(logo_path):
                try:
                    logo = Image(logo_path, width=2.5 * inch, height=2.5 * inch)
                    logo.hAlign = 'CENTER'
                    story.append(logo)
                    story.append(Spacer(1, 0.5 * inch))
                except:
                    pass  # Skip logo if there's an error loading it

            # Index name
            story.append(Paragraph(metadata['index_name'], title_style))
            story.append(Spacer(1, 0.3 * inch))

            # Generated date
            story.append(Paragraph(f"<i>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>",
                                 ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=10,
                                              alignment=TA_CENTER, textColor=colors.HexColor('#666666'))))

            # Move to column 2 for Books and Properties
            story.append(FrameBreak())

            # Column 2-3-4: Books section (can overflow across columns)
            if metadata['books']:
                story.append(Paragraph("Books", front_heading_style))
                for book in metadata['books']:
                    page_info = f" ({book['page_count']} pages)" if book['page_count'] else ""
                    story.append(Paragraph(f"<b>Book {book['book_number']}:</b> {self.escape_html(book['book_name'])}{page_info}",
                                         front_content_style))
                story.append(Spacer(1, 0.3 * inch))

            # Custom properties section (follows books, can overflow)
            if metadata['custom_properties']:
                story.append(Paragraph("Properties", front_heading_style))
                for prop in metadata['custom_properties']:
                    story.append(Paragraph(f"<b>{self.escape_html(prop['name'])}:</b> {self.escape_html(prop['value'])}",
                                         front_content_style))
                story.append(Spacer(1, 0.3 * inch))

            # Page break before index
            story.append(PageBreak())

            # Now add index entries
            current_letter = None

            for term, references in entries:
                # Get first letter (normalize special chars and numbers to "#")
                first_letter = self.normalize_first_letter(term)

                # Add letter heading when we encounter a new letter
                if first_letter != current_letter:
                    if current_letter is not None:
                        story.append(Spacer(1, 0.1 * inch))

                    # Add letter heading with grey background
                    heading_para = Paragraph(f"<b>{first_letter}</b>", heading_style)
                    heading_table = Table([[heading_para]], colWidths=[frame_width - 12])
                    heading_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f1f5f9')),
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    story.append(heading_table)
                    current_letter = first_letter

                # Format entry: term in bold, references in normal text
                ref_str = ", ".join(references)
                entry_text = f"<b>{self.escape_html(term)}</b>: {self.escape_html(ref_str)}"
                story.append(Paragraph(entry_text, entry_style))

            # Build PDF
            doc.build(story)
            return True

        except ImportError:
            return False

    @staticmethod
    def escape_html(text: str) -> str:
        """Escape HTML special characters for use in ReportLab"""
        replacements = {
            '&': '&amp;',
            '<': '&lt;',
            '>': '&gt;',
        }

        for char, replacement in replacements.items():
            text = text.replace(char, replacement)

        return text

    @staticmethod
    def escape_latex(text: str) -> str:
        """Escape special LaTeX characters"""
        replacements = {
            '&': r'\&',
            '%': r'\%',
            '$': r'\$',
            '#': r'\#',
            '_': r'\_',
            '{': r'\{',
            '}': r'\}',
            '~': r'\textasciitilde{}',
            '^': r'\textasciicircum{}',
            '\\': r'\textbackslash{}',
        }

        for char, replacement in replacements.items():
            text = text.replace(char, replacement)

        return text

    def format_notes_text(self, notes: List[Tuple[str, str]]) -> str:
        """
        Format notes as plain text
        """
        if not notes:
            return "Empty notes\n"

        # Sort notes with "#" (special chars/numbers) at the end
        def sort_key(item):
            term = item[0]
            first_letter = self.normalize_first_letter(term)
            # Put "#" at beginning, otherwise sort alphabetically
            return (first_letter != '#', term.lower())

        notes = sorted(notes, key=sort_key)

        output = []

        # Add header
        output.append("=" * 60)
        output.append("Notes".center(60))
        output.append("=" * 60)
        output.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        output.append("=" * 60)
        output.append("")

        for term, note in notes:
            output.append(f"{term}")
            output.append("-" * 40)
            output.append(note)
            output.append("")

        output.append("=" * 60)
        output.append(f"Total notes: {len(notes)}")
        output.append("=" * 60)

        return "\n".join(output)

    def format_notes_csv(self, notes: List[Tuple[str, str]]) -> str:
        """
        Format notes as CSV
        """
        import csv
        from io import StringIO

        # Sort notes with "#" (special chars/numbers) at the end
        def sort_key(item):
            term = item[0]
            first_letter = self.normalize_first_letter(term)
            # Put "#" at beginning, otherwise sort alphabetically
            return (first_letter != '#', term.lower())

        notes = sorted(notes, key=sort_key)

        output = StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(['Term', 'Notes'])

        # Write data
        for term, note in notes:
            writer.writerow([term, note])

        return output.getvalue()

    def format_notes_pdf(self, notes: List[Tuple[str, str]], output_path: str) -> bool:
        """
        Format notes as a 2-column landscape PDF
        Returns True if successful
        """
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus.doctemplate import BaseDocTemplate, PageTemplate
            from reportlab.platypus.frames import Frame

            if not notes:
                return False

            # Sort notes with "#" (special chars/numbers) at the end
            def sort_key(item):
                term = item[0]
                first_letter = self.normalize_first_letter(term)
                # Put "#" at end, otherwise sort alphabetically
                return (first_letter == '#', term.lower())

            notes = sorted(notes, key=sort_key)

            # Define page number footer function
            def add_page_number(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica', 9)
                page_num = canvas.getPageNumber()
                text = f"{page_num}"
                canvas.drawRightString(landscape(letter)[0] - 0.5 * inch, 0.3 * inch, text)
                canvas.restoreState()

            # Create PDF with landscape orientation
            doc = BaseDocTemplate(
                output_path,
                pagesize=landscape(letter),
                rightMargin=0.5 * inch,
                leftMargin=0.5 * inch,
                topMargin=0.75 * inch,
                bottomMargin=0.5 * inch
            )

            # Calculate frame dimensions for 2 columns
            page_width, page_height = landscape(letter)
            frame_width = (page_width - 1 * inch - 0.3 * inch) / 2  # subtract margins and gap
            frame_height = page_height - 1.25 * inch  # subtract top and bottom margins

            # Create 2 frames (columns)
            frames = []
            for i in range(2):
                x = 0.5 * inch + i * (frame_width + 0.3 * inch)
                frame = Frame(
                    x, 0.5 * inch,
                    frame_width, frame_height,
                    leftPadding=10, rightPadding=10,
                    topPadding=10, bottomPadding=10,
                    showBoundary=0
                )
                frames.append(frame)

            # Create page template with 2 columns and page numbers
            template = PageTemplate(id='TwoCol', frames=frames, onPage=add_page_number)
            doc.addPageTemplates([template])

            # Styles
            styles = getSampleStyleSheet()

            # Term heading style
            term_style = ParagraphStyle(
                'TermStyle',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#2563eb'),
                spaceAfter=6,
                spaceBefore=10,
                fontName='Helvetica-Bold',
                keepWithNext=True
            )

            # Notes text style
            notes_style = ParagraphStyle(
                'NotesStyle',
                parent=styles['Normal'],
                fontSize=9,
                leftIndent=10,
                spaceAfter=12,
                leading=12
            )

            # Build content
            story = []

            for term, note in notes:
                # Add term heading
                story.append(Paragraph(f"<b>{self.escape_html(term)}</b>", term_style))

                # Add notes text
                # Preserve line breaks in notes
                note_html = self.escape_html(note).replace('\n', '<br/>')
                story.append(Paragraph(note_html, notes_style))

            # Build PDF
            doc.build(story)
            return True

        except ImportError:
            return False

    def format_csv(self, entries: List[Tuple[str, List[str]]],
                   metadata: dict = None) -> str:
        """
        Format index entries as CSV
        """
        import csv
        from io import StringIO

        if metadata is None:
            metadata = {'index_name': 'Index', 'books': [], 'custom_properties': []}

        # Sort entries
        def sort_key(item):
            term = item[0]
            first_letter = self.normalize_first_letter(term)
            return (first_letter != '#', term.lower())

        entries = sorted(entries, key=sort_key)

        output = StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(['Term', 'References'])

        # Write data
        for term, references in entries:
            ref_str = ", ".join(references)
            writer.writerow([term, ref_str])

        return output.getvalue()

    def format_excel(self, entries: List[Tuple[str, List[str]]],
                     output_path: str, metadata: dict = None) -> bool:
        """
        Format index entries as an Excel file
        Returns True if successful
        """
        if metadata is None:
            metadata = {'index_name': 'Index', 'books': [], 'custom_properties': []}

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            if not entries:
                return False

            # Sort entries
            def sort_key(item):
                term = item[0]
                first_letter = self.normalize_first_letter(term)
                return (first_letter == '#', term.lower())

            entries = sorted(entries, key=sort_key)

            # Create workbook
            wb = Workbook()

            # Metadata sheet
            ws_meta = wb.active
            ws_meta.title = "Index Info"

            # Add index name
            ws_meta['A1'] = metadata['index_name']
            ws_meta['A1'].font = Font(size=16, bold=True)
            ws_meta['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # Add books
            if metadata['books']:
                row = 4
                ws_meta[f'A{row}'] = "Books"
                ws_meta[f'A{row}'].font = Font(size=12, bold=True)
                row += 1
                for book in metadata['books']:
                    page_info = f" ({book['page_count']} pages)" if book['page_count'] else ""
                    ws_meta[f'A{row}'] = f"Book {book['book_number']}: {book['book_name']}{page_info}"
                    row += 1
                row += 1

            # Add custom properties
            if metadata['custom_properties']:
                if not metadata['books']:
                    row = 4
                ws_meta[f'A{row}'] = "Properties"
                ws_meta[f'A{row}'].font = Font(size=12, bold=True)
                row += 1
                for prop in metadata['custom_properties']:
                    ws_meta[f'A{row}'] = f"{prop['name']}: {prop['value']}"
                    row += 1

            # Adjust column width
            ws_meta.column_dimensions['A'].width = 60

            # Index entries sheet
            ws = wb.create_sheet("Index")

            # Headers
            ws['A1'] = "Term"
            ws['B1'] = "References"
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['A1'].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            ws['B1'].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

            # Add entries
            row = 2
            current_letter = None

            for term, references in entries:
                first_letter = self.normalize_first_letter(term)

                # Add letter heading
                if first_letter != current_letter:
                    ws[f'A{row}'] = first_letter
                    ws[f'A{row}'].font = Font(size=14, bold=True)
                    ws[f'A{row}'].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                    ws.merge_cells(f'A{row}:B{row}')
                    row += 1
                    current_letter = first_letter

                # Add entry
                ws[f'A{row}'] = term
                ws[f'B{row}'] = ", ".join(references)
                row += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 60

            # Save workbook
            wb.save(output_path)
            return True

        except ImportError:
            return False

    def format_notes_latex(self, notes: List[Tuple[str, str]]) -> str:
        """
        Format notes as LaTeX
        """
        if not notes:
            return "% Empty notes\n"

        # Sort notes
        def sort_key(item):
            term = item[0]
            first_letter = self.normalize_first_letter(term)
            return (first_letter != '#', term.lower())

        notes = sorted(notes, key=sort_key)

        output = []

        # Header
        output.append("% Notes")
        output.append(f"% Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        output.append("")
        output.append("\\documentclass{article}")
        output.append("\\usepackage[utf8]{inputenc}")
        output.append("\\title{Notes}")
        output.append("\\date{" + datetime.now().strftime('%Y-%m-%d') + "}")
        output.append("")
        output.append("\\begin{document}")
        output.append("\\maketitle")
        output.append("")

        for term, note in notes:
            output.append(f"\\section*{{{self.escape_latex(term)}}}")
            output.append("")
            # Escape note content and preserve line breaks
            escaped_note = self.escape_latex(note)
            # Replace double newlines with paragraph breaks
            escaped_note = escaped_note.replace('\n\n', '\n\n\\par\n')
            output.append(escaped_note)
            output.append("")

        output.append("\\end{document}")

        return "\n".join(output)

    def format_notes_markdown(self, notes: List[Tuple[str, str]]) -> str:
        """
        Format notes as Markdown
        """
        if not notes:
            return "*Empty notes*\n"

        # Sort notes
        def sort_key(item):
            term = item[0]
            first_letter = self.normalize_first_letter(term)
            return (first_letter != '#', term.lower())

        notes = sorted(notes, key=sort_key)

        output = []

        # Header
        output.append("# Notes")
        output.append("")
        output.append(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
        output.append("")
        output.append("---")
        output.append("")

        for term, note in notes:
            output.append(f"## {term}")
            output.append("")
            output.append(note)
            output.append("")

        output.append("---")
        output.append(f"*Total notes: {len(notes)}*")

        return "\n".join(output)

    def format_notes_excel(self, notes: List[Tuple[str, str]], output_path: str) -> bool:
        """
        Format notes as an Excel file
        Returns True if successful
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            if not notes:
                return False

            # Sort notes
            def sort_key(item):
                term = item[0]
                first_letter = self.normalize_first_letter(term)
                return (first_letter == '#', term.lower())

            notes = sorted(notes, key=sort_key)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Notes"

            # Headers
            ws['A1'] = "Term"
            ws['B1'] = "Notes"
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            ws['A1'].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            ws['B1'].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

            # Add notes
            row = 2
            for term, note in notes:
                ws[f'A{row}'] = term
                ws[f'B{row}'] = note
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                row += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 80

            # Save workbook
            wb.save(output_path)
            return True

        except ImportError:
            return False